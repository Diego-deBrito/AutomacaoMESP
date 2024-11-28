import os
import pyperclip
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from datetime import datetime, timedelta
import pandas as pd
import chromedriver_autoinstaller
import time


# Configuração inicial do ChromeDriver
def conectar_navegador_existente():
    """Conecta ao navegador Chrome já aberto, utilizando a porta de depuração 9222."""
    try:
        chromedriver_autoinstaller.install()
        options = webdriver.ChromeOptions()
        options.debugger_address = "localhost:9222"
        driver = webdriver.Chrome(options=options)
        return driver
    except Exception as e:
        print(f"Erro ao conectar ao navegador existente: {e}")
        return None


def clicar_elemento(driver, xpath, tempo_espera=10):
    """Espera e clica em um elemento identificado pelo XPath."""
    try:
        elemento = WebDriverWait(driver, tempo_espera).until(
            EC.element_to_be_clickable((By.XPATH, xpath))
        )
        elemento.click()
        print(f"Elemento clicado: {xpath}")
    except (TimeoutException, NoSuchElementException) as e:
        print(f"Erro ao encontrar ou clicar no elemento {xpath}: {e}")
    except Exception as e:
        print(f"Erro inesperado: {e}")


def coletar_dados_instrumentos_pandas(nome_arquivo, aba):
    """Extrai dados usando pandas para garantir a leitura correta das colunas."""
    try:
        planilha = pd.read_excel(nome_arquivo, sheet_name=aba)
        dados_filtrados = planilha[
            (planilha["Status"] == "ATIVOS TODOS") &
            (planilha["Instrumento nº"].notna()) &
            (planilha["Técnico"].notna()) &
            (planilha["e-mail do Técnico"].notna())
            ][["Instrumento nº", "Técnico", "e-mail do Técnico"]]

        dados_filtrados["e-mail do Técnico"].fillna("não encontrado", inplace=True)
        dados_filtrados["Instrumento nº"] = dados_filtrados["Instrumento nº"].apply(
            lambda x: str(int(x)) if pd.notna(x) else "não encontrado"
        )

        dados_instrumentos = dados_filtrados.to_dict(orient="records")
        print(f"Instrumentos coletados: {len(dados_instrumentos)}")
        return dados_instrumentos
    except Exception as e:
        print(f"Erro ao coletar dados do Excel: {e}")
        return []


def extrair_data_termino(driver):
    """Extrai a data de término de vigência usando o XPath."""
    try:
        xpath_data_termino = '//*[@id="tr-alterarTerminoVigencia"]/td[2]'
        elemento_data = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, xpath_data_termino))
        )
        data_texto = elemento_data.text.strip()
        print(f"Data extraída: {data_texto}")
        return datetime.strptime(data_texto, "%d/%m/%Y")
    except Exception as e:
        print(f"Erro ao extrair data de término: {e}")
        return None


def gerar_planilha_incremental(instrumento):
    """Atualiza a planilha Excel incrementando com um novo instrumento a cada chamada."""
    nome_arquivo_excel = "C:/Users/diego.brito/Downloads/robov1/Resultados_Instrumentos.xlsx"

    try:
        # Verifica se o arquivo já existe
        if os.path.exists(nome_arquivo_excel):
            workbook = load_workbook(nome_arquivo_excel)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Instrumentos"
            sheet.append([
                "Instrumento nº", "Data de Término", "Modalidade",
                "Data de Notificação 1", "Data de Notificação 2",
                "Notificação Enviada", "Técnico", "E-mail"
            ])
            print(f"Arquivo criado: {nome_arquivo_excel}")

        # Adiciona nova linha com os dados do instrumento
        nova_linha = [
            instrumento.get("Instrumento nº", "N/A"),
            instrumento.get("Data de Término", "N/A"),
            instrumento.get("Modalidade", "N/A"),
            instrumento.get("Data de Notificação 1", "N/A"),
            instrumento.get("Data de Notificação 2", "N/A"),
            instrumento.get("Notificação Enviada", "N/A"),
            instrumento.get("Técnico", "N/A"),
            instrumento.get("e-mail do Técnico", "N/A")
        ]
        sheet.append(nova_linha)
        workbook.save(nome_arquivo_excel)
        print(f"Dados salvos no Excel: {nova_linha}")
    except Exception as e:
        print(f"Erro ao atualizar o Excel: {e}")


def executar_processo():
    """Fluxo principal do programa."""
    driver = conectar_navegador_existente()
    if not driver:
        print("Não foi possível conectar ao navegador.")
        return

    dados_instrumentos = coletar_dados_instrumentos_pandas(
        "CONTROLE DE PARCERIAS CGAP.xlsx", "PARCERIAS CGAP"
    )
    if not dados_instrumentos:
        print("Nenhum dado encontrado para processamento.")
        return

    for instrumento in dados_instrumentos:
        try:
            numero_instrumento = instrumento.get("Instrumento nº")
            print(f"Processando instrumento: {numero_instrumento}")

            clicar_elemento(driver, '//*[@id="menuPrincipal"]/div[1]/div[4]')
            clicar_elemento(driver, '//*[@id="contentMenu"]/div[1]/ul/li[6]/a')

            # Pesquisar pelo número do instrumento
            campo_pesquisa_xpath = '//*[@id="consultarNumeroConvenio"]'
            campo = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, campo_pesquisa_xpath))
            )
            campo.clear()
            campo.send_keys(str(numero_instrumento))

            botao_submit_xpath = '//*[@id="form_submit"]'
            clicar_elemento(driver, botao_submit_xpath)

            time.sleep(2)

            detalhe_instrumento_xpath = '//*[@id="instrumentoId"]/a'
            clicar_elemento(driver, detalhe_instrumento_xpath)

            # Extrai informações do instrumento
            data_termino = extrair_data_termino(driver)
            if not data_termino:
                continue

            instrumento["Data de Término"] = data_termino.strftime("%d/%m/%Y")
            instrumento["Modalidade"] = "Exemplo"  # Substituir por extração real
            gerar_planilha_incremental(instrumento)

        except Exception as e:
            print(f"Erro ao processar instrumento {numero_instrumento}: {e}")
            continue


if __name__ == "__main__":
    executar_processo()
