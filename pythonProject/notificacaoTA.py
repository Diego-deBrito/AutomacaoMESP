import os
import pyperclip
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from datetime import datetime, timedelta
import pandas as pd
from openpyxl.styles import PatternFill
import chromedriver_autoinstaller
import time

# Configuração inicial do ChromeDriver
def conectar_navegador_existente():
    """Conecta ao navegador Chrome já aberto, utilizando a porta de depuração 9222."""
    try:
        chromedriver_autoinstaller.install()
        options = webdriver.ChromeOptions()
        options.debugger_address = "localhost:9222"
        driver = webdriver.Chrome(options=options)
        return driver
    except Exception as e:
        print(f"Erro ao conectar ao navegador existente: {e}")
        return None


def clicar_elemento(driver, xpath, tempo_espera=10):
    """Espera e clica em um elemento identificado pelo XPath."""
    try:
        elemento = WebDriverWait(driver, tempo_espera).until(
            EC.element_to_be_clickable((By.XPATH, xpath))
        )
        elemento.click()
        print(f"Elemento clicado: {xpath}")
    except (TimeoutException, NoSuchElementException) as e:
        print(f"Erro ao encontrar ou clicar no elemento {xpath}: {e}")
    except Exception as e:
        print(f"Erro inesperado: {e}")


def coletar_dados_instrumentos_pandas(nome_arquivo, aba):
    """Extrai dados usando pandas para garantir a leitura correta das colunas."""
    try:
        planilha = pd.read_excel(nome_arquivo, sheet_name=aba)
        dados_filtrados = planilha[
            (planilha["Status"] == "ATIVOS TODOS") &
            (planilha["Instrumento nº"].notna()) &
            (planilha["Técnico"].notna()) &
            (planilha["e-mail do Técnico"].notna())
        ][["Instrumento nº", "Técnico", "e-mail do Técnico"]]
        dados_filtrados["e-mail do Técnico"].fillna("não encontrado", inplace=True)
        dados_filtrados["Instrumento nº"] = dados_filtrados["Instrumento nº"].apply(lambda x: str(int(x)) if pd.notna(x) else "não encontrado")
        dados_instrumentos = dados_filtrados.to_dict(orient="records")
        for item in dados_instrumentos:
            print(f"Instrumento: {item['Instrumento nº']}, Técnico: {item['Técnico']}, E-mail: {item['e-mail do Técnico']}")
        return dados_instrumentos
    except Exception as e:
        print(f"Erro ao coletar dados do Excel com pandas: {e}")
        return []


def extrair_data_termino(driver):
    """Extrai a data de término de vigência usando o XPath."""
    try:
        xpath_data_termino = '//*[@id="tr-alterarTerminoVigencia"]/td[2]'
        elemento_data = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, xpath_data_termino))
        )
        data_texto = elemento_data.text.strip()
        print(f"Data extraída: {data_texto}")
        data_termino = datetime.strptime(data_texto, "%d/%m/%Y")
        return data_termino
    except Exception as e:
        print(f"Erro ao extrair data de término: {e}")
        return None



def gerar_planilha_incremental(instrumento):
    """
    Atualiza ou cria uma planilha Excel contendo as informações do instrumento processado.
    Adiciona colunas com as seguintes informações:
    - Instrumento nº, Data de Término, Notificação 1, Notificação 2, Técnico, Email do Técnico
    """
    # Define o nome do arquivo Excel
    nome_arquivo_excel = r"C:/Temp/Instrumentos_Parcerias.xlsx"

    # Garante que o diretório existe; se não, cria o diretório
    diretorio = os.path.dirname(nome_arquivo_excel)
    if not os.path.exists(diretorio):
        try:
            os.makedirs(diretorio)
            print(f"Diretório criado: {diretorio}")
        except Exception as e:
            print(f"Erro ao criar o diretório {diretorio}: {e}")
            return

    # Verifica se o arquivo já existe
    if not os.path.exists(nome_arquivo_excel):
        # Se o arquivo não existir, cria o arquivo com cabeçalhos
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Instrumentos"
        # Adiciona cabeçalhos
        sheet.append([
            "Instrumento nº",
            "Data de Término da Vigência",
            "Notificação 1",
            "Notificação 2",
            "Técnico",
            "Email do Técnico"
        ])
        workbook.save(nome_arquivo_excel)
        print(f"Arquivo '{nome_arquivo_excel}' criado com cabeçalhos.")

    # Carrega a planilha existente
    workbook = load_workbook(nome_arquivo_excel)
    sheet = workbook["Instrumentos"]

    # Adiciona os dados do instrumento na próxima linha
    nova_linha = [
        instrumento.get("Instrumento nº"),
        instrumento.get("Data de Término"),
        instrumento.get("Data de Notificação 1"),
        instrumento.get("Data de Notificação 2"),
        instrumento.get("Técnico"),
        instrumento.get("Email do Técnico"),
    ]
    sheet.append(nova_linha)

    # Salva a planilha
    workbook.save(nome_arquivo_excel)
    print(f"Instrumento adicionado ao arquivo '{nome_arquivo_excel}': {nova_linha}")





def calcular_notificacoes(modalidade, data_termino):
    """
    Calcula as datas de notificação com base na modalidade e na data de término.

    Args:
        modalidade (str): A modalidade do instrumento (Ex.: "Termo de Fomento", "Convênio").
        data_termino (datetime): A data de término da vigência do instrumento.

    Returns:
        list[datetime]: Uma lista com as datas de notificação calculadas.
    """
    try:
        if modalidade == "Termo de Fomento":
            # Notificar com 60 e 45 dias antes do término
            notificacoes = [
                data_termino - timedelta(days=60),
                data_termino - timedelta(days=45),
            ]
        elif modalidade == "Convênio":
            # Notificar com 90 e 75 dias antes do término
            notificacoes = [
                data_termino - timedelta(days=90),
                data_termino - timedelta(days=75),
            ]
        else:
            print(f"Modalidade desconhecida: {modalidade}. Não será possível calcular as notificações.")
            notificacoes = []

        return notificacoes

    except Exception as e:
        print(f"Erro ao calcular notificações para modalidade '{modalidade}': {e}")
        return []



def executar_processo():
    """
    Fluxo principal:
    - Verifica os instrumentos com base nas regras de notificação.
    - Cria e atualiza a planilha conforme os instrumentos são processados.
    """
    driver = conectar_navegador_existente()
    if driver:
        # Coleta os dados dos instrumentos do Excel
        try:
            dados_instrumentos = coletar_dados_instrumentos_pandas(
                r'C:/Users/d-deb/OneDrive/Documents/dev/robov1/CONTROLE DE PARCERIAS CGAP.xlsx', "PARCERIAS CGAP"
            )
            print("Dados dos instrumentos carregados:", dados_instrumentos)

        except Exception as e:
            print(f"Erro ao carregar os dados do Excel: {e}")
            return

        for idx, instrumento in enumerate(dados_instrumentos):
            try:
                numero_instrumento = instrumento.get("Instrumento nº")
                tecnico = instrumento.get("Técnico")
                email_tecnico = instrumento.get("e-mail do Técnico")

                if not isinstance(numero_instrumento, (int, str)):
                    print(f"Formato inesperado para o número do instrumento: {numero_instrumento}")
                    continue

                # Navega para a página do instrumento
                clicar_elemento(driver, '//*[@id="menuPrincipal"]/div[1]/div[4]')
                clicar_elemento(driver, '//*[@id="contentMenu"]/div[1]/ul/li[6]/a')

                # Localiza o campo de pesquisa e o botão de submissão
                campo_pesquisa_xpath = '//*[@id="consultarNumeroConvenio"]'
                botao_submit_xpath = '//*[@id="form_submit"]'

                campo = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, campo_pesquisa_xpath))
                )
                campo.clear()
                campo.send_keys(str(numero_instrumento))

                # Clica no botão de submissão
                submit_button = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, botao_submit_xpath))
                )
                submit_button.click()

                time.sleep(2)

                detalhe_instrumento_xpath = '//*[@id="instrumentoId"]/a'
                elemento_detalhe = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, detalhe_instrumento_xpath))
                )
                elemento_detalhe.click()

                # Extrai a data de término de vigência
                data_termino = extrair_data_termino(driver)
                if not data_termino:
                    print(f"Data de término não encontrada para o instrumento {numero_instrumento}")
                    clicar_elemento(driver, '//*[@id="logo"]/a')  # Retorna à página inicial
                    continue

                # Extrai a modalidade
                modalidade_xpath = '//*[@id="tr-alterarModalidade"]/td[2]/table/tbody/tr/td[1]'
                elemento_modalidade = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, modalidade_xpath))
                )
                modalidade = elemento_modalidade.text.strip()
                print(f"Modalidade extraída para o instrumento {numero_instrumento}: {modalidade}")

                # Calcula as notificações
                notificacoes = calcular_notificacoes(modalidade, data_termino)
                if len(notificacoes) != 2:
                    print(f"Erro ao calcular as notificações para o instrumento {numero_instrumento}.")
                    clicar_elemento(driver, '//*[@id="logo"]/a')  # Retorna à página inicial
                    continue

                # Registra as informações na planilha
                gerar_planilha_incremental({
                    "Instrumento nº": numero_instrumento,
                    "Data de Término": data_termino.strftime("%d/%m/%Y"),
                    "Data de Notificação 1": notificacoes[0].strftime("%d/%m/%Y"),
                    "Data de Notificação 2": notificacoes[1].strftime("%d/%m/%Y"),
                    "Técnico": tecnico,
                    "Email do Técnico": email_tecnico,
                })

                print(f"Dados registrados para o instrumento {numero_instrumento}.")

                # Retorna à página inicial antes de continuar
                clicar_elemento(driver, '//*[@id="logo"]/a')
                time.sleep(2)

            except Exception as e:
                print(f"Erro ao processar o instrumento {numero_instrumento}: {e}")
                clicar_elemento(driver, '//*[@id="logo"]/a')  # Retorna à página inicial em caso de erro
                continue

    else:
        print("Não foi possível conectar ao navegador. Encerrando o processo.")



if __name__ == "__main__":
    executar_processo()
