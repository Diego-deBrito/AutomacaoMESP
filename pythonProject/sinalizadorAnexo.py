import os
import openpyxl
import pyperclip
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime

from AjustePT import clicar_elemento
from robov1 import clicar_e_colar


def conectar_navegador_existente(retentativas=3):
    """Conecta ao navegador Chrome já aberto, utilizando a porta de depuração 9222, com múltiplas tentativas."""
    for tentativa in range(1, retentativas + 1):
        try:
            print(f"[INFO] Tentativa {tentativa} de conectar ao navegador na porta 9222...")
            options = webdriver.ChromeOptions()
            options.debugger_address = "localhost:9222"
            driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
            print("[INFO] Conectado ao navegador existente com sucesso.")
            return driver
        except WebDriverException as e:
            print(f"[ERRO] Erro ao conectar ao navegador (tentativa {tentativa}): {e}")
            time.sleep(3)
    print("[ERRO] Não foi possível conectar ao navegador após múltiplas tentativas.")
    return None


def acessar_aba_anexos(driver):
    """Navega até a aba de anexos."""
    try:
        clicar_elemento(driver, '//*[@id="div_997366806"]/span/span')  # Menu principal
        clicar_elemento(driver, '//*[@id="menu_link_997366806_1965609892"]/div/span/span')  # Submenu
        clicar_elemento(driver, '/html/body/div[3]/div[15]/div[3]/div[1]/div/form/table/tbody/tr/td[2]/input[2]')  # Seleciona o instrumento
    except Exception as e:
        print(f"[ERRO] Erro ao acessar a aba de anexos: {e}")
        raise


def coletar_dados_instrumentos(nome_arquivo, aba):
    """Extrai dados das colunas 'Número do Instrumento', 'Técnico Responsável' e 'Email' apenas para instrumentos ativos."""
    try:
        wb = openpyxl.load_workbook(nome_arquivo)
        ws = wb[aba]
        cabecalho = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        print(f"[DEBUG] Cabeçalho encontrado: {cabecalho}")

        required_columns = ["Instrumento nº", "Técnico", "e-mail do Técnico", "Status"]
        col_indices = {col: cabecalho.index(col) for col in required_columns if col in cabecalho}

        if len(col_indices) < len(required_columns):
            print("[ERRO] Algumas colunas obrigatórias não foram encontradas no Excel de controle.")
            return []

        dados_instrumentos = [
            [
                row[col_indices["Instrumento nº"]].value or "Não encontrado",
                row[col_indices["Técnico"]].value or "Não encontrado",
                row[col_indices["e-mail do Técnico"]].value or "Não encontrado",
            ]
            for row in ws.iter_rows(min_row=2)
            if row[col_indices["Status"]].value == "ATIVOS TODOS"
        ]
        print(f"[INFO] {len(dados_instrumentos)} instrumentos ativos encontrados.")
        return dados_instrumentos
    except Exception as e:
        print(f"[ERRO] Erro ao coletar dados do Excel: {e}")
        return []


def inicializar_arquivo_excel(caminho_arquivo_excel):
    """Inicializa o arquivo Excel e cria o cabeçalho se ele ainda não existir."""
    diretorio_excel = os.path.dirname(caminho_arquivo_excel)
    if not os.path.exists(diretorio_excel):
        os.makedirs(diretorio_excel)

    if not os.path.exists(caminho_arquivo_excel):
        workbook = openpyxl.Workbook()
        planilha = workbook.active
        planilha.title = "Dados Processados"
        planilha.append(["Número do Instrumento", "Técnico Responsável", "Email", "AnexosExistentes", "NovosAnexos"])
        workbook.save(caminho_arquivo_excel)
        print(f"[INFO] Arquivo Excel '{caminho_arquivo_excel}' inicializado com o cabeçalho.")
    else:
        print(f"[INFO] Arquivo Excel '{caminho_arquivo_excel}' já existe e será atualizado.")


def capturar_data_ultimo_anexo(driver, numero_do_instrumento):
    """Captura a data de upload mais atual dos anexos na tabela de anexos para um instrumento específico."""
    try:
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="tbodyrow"]/tr')))
        data_elements = driver.find_elements(By.XPATH, '//*[@id="tbodyrow"]/tr/td[3]/div')
        datas = [
            datetime.strptime(element.text.strip(), "%d/%m/%Y")
            for element in data_elements if element.text.strip()
        ]
        if datas:
            ultima_data = max(datas)
            return ultima_data.strftime("%d/%m/%Y")
        else:
            return "Sem anexos"
    except (TimeoutException, NoSuchElementException, ValueError):
        return "Erro ao capturar data"


def executar_processo():
    """Executa o processo principal de coleta de dados e atualização no Excel."""
    navegador_web = conectar_navegador_existente()
    if not navegador_web:
        return

    # Caminhos dos arquivos
    caminho_arquivo_controle = r'C:\Users\diego.brito\Downloads\robov1\CONTROLE DE PARCERIAS CGAP.xlsx'
    caminho_novo_arquivo_excel = r'C:\Users\diego.brito\Downloads\robov1\relatorio_instrumentos.xlsx'
    nome_da_aba_controle = 'PARCERIAS CGAP'

    # Inicializa o arquivo Excel de saída
    inicializar_arquivo_excel(caminho_novo_arquivo_excel)

    # Coleta os dados dos instrumentos no arquivo de controle
    lista_dados_instrumentos = coletar_dados_instrumentos(caminho_arquivo_controle, nome_da_aba_controle)

    if not lista_dados_instrumentos:
        print("[ERRO] Nenhum dado foi coletado. Processo encerrado.")
        return

    dados_para_salvar = []
    total_instrumentos = len(lista_dados_instrumentos)

    for idx, (numero_do_instrumento, tecnico_responsavel, email_tecnico) in enumerate(lista_dados_instrumentos, start=1):
        inicio = time.perf_counter()
        print(f"[INFO] Processando instrumento {idx}/{total_instrumentos}: {numero_do_instrumento}")
        try:
            clicar_elemento(navegador_web, '//*[@id="logo"]/a')
            pyperclip.copy(str(numero_do_instrumento))
            clicar_elemento(navegador_web, '//*[@id="menuPrincipal"]/div[1]/div[4]')
            clicar_elemento(navegador_web, '//*[@id="contentMenu"]/div[1]/ul/li[6]/a')
            WebDriverWait(navegador_web, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="consultarNumeroConvenio"]'))
            ).send_keys(numero_do_instrumento)
            clicar_elemento(navegador_web, '//*[@id="form_submit"]')
            clicar_elemento(navegador_web, '//*[@id="instrumentoId"]/a')

            acessar_aba_anexos(navegador_web)
            data_anexo = capturar_data_ultimo_anexo(navegador_web, numero_do_instrumento)

            dados_para_salvar.append([
                numero_do_instrumento,
                tecnico_responsavel,
                email_tecnico,
                data_anexo,
                "Nenhum",
            ])
        except Exception as erro:
            print(f"[ERRO] Instrumento {idx}/{total_instrumentos}: Falha ao processar {numero_do_instrumento}: {erro}")
        finally:
            fim = time.perf_counter()
            print(f"[INFO] Instrumento {idx}/{total_instrumentos} processado em {fim - inicio:.2f} segundos.")

    # Salvar todos os dados no Excel ao final
    try:
        workbook = openpyxl.load_workbook(caminho_novo_arquivo_excel)
        planilha = workbook.active
        for linha in dados_para_salvar:
            print(f"[DEBUG] Salvando linha: {linha}")
            planilha.append(linha)
        workbook.save(caminho_novo_arquivo_excel)
        print("[INFO] Dados salvos no Excel com sucesso.")
    except Exception as e:
        print(f"[ERRO] Falha ao salvar os dados no Excel: {e}")


if __name__ == "__main__":
    executar_processo()
